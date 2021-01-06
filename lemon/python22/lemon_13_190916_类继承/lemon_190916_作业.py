# --*-- coding : utf-8 --*--
# Project      : python22
# Current file : lemon_190916_作业.py
# Author       : Administrator
# Create time  : 2019-09-16 21:02
# IDE          : PyCharm
# TODO 成长很苦, 进步很甜, 加油！

# 0916-作业练习手感
# 截止时间：09月18日14:00
# 1、在之前定义的手机类下面定义智能手机类和苹果手机类。
# 智能手机打电话具有录音功能。 苹果手机打电话不仅具有录音功能，还有 facetime 功能


class Phone:
    # name = name  这是类属性：类属性是写在初始化方法外，类下方的

    # __init__ 初始化方法，初始化函数
    def __init__(self, system, brand, versions):
        """ 实例属性是写在初始化方法里面的 """
        self.system = system  # 系统
        self.brand = brand  # 品牌
        self.versions = versions  # 型号

    def ring_up(self, family, *args):
        """ 打电话，调用实例属性 """
        return f'用{self.system}{self.brand}手机给{family, args}打电话'

    def send_messages(self, family, *args):
        """ 发短信，调用实例属性 """
        return f'用{self.brand}{self.versions}手机给{family, args}发短信'

    def look_movie(self, family, *args):
        """ 看电影，调用实例属性 """
        return f'用{self.brand}{self.versions}手机和{family, args}看电影'


class Iphone(Phone):
    """苹果手机"""
    system = '苹果'

    def record(self, family):
        return f'{Iphone.system}手机还能录音, 听{family}唱的戏'


class SmartPhone(Iphone):
    """智能手机"""
    system = '智能手机'
    facetime = '视频聊天'

    def __init__(self, system, brand, versions):
        super().__init__(system, brand, versions)

    def facetime_01(self, girlfriend):
        return f'{SmartPhone.system}还可以和{girlfriend}{SmartPhone.facetime}'


# '安卓', '华为', 'P40'
# 智能手机类
if __name__ == '__main__':
    phone = SmartPhone('安卓', '华为', 'P40')  # 实例化对象
    print(phone.ring_up('老妈', '老爸', '奶奶'))  # 打电话
    print(phone.send_messages('老妈', '老爸', '奶奶'))  # 发短信
    print(phone.look_movie('老妈', '老爸', '奶奶'))  # 看电影
    print(phone.record('爷爷'))  # 录音
    print(phone.facetime_01('女朋友'))  # 视频通话

# 2、
# 定义一个 ExcelManual 类。具有获取 sheet 表单， 读取单元格 和 修改单元格功能。

import openpyxl


# from openpyxl import Workbook


class ExcelManual:

    def found_file(self):
        """ 创建Workbook, 并且保存 """''
        wb = openpyxl.Workbook()
        wb.save('sheet.xlsx')
        print('创建文件成功')

    def examine(self):
        """ 写入表格 """''
        # 打开已有的excel
        wb = openpyxl.load_workbook('sheet.xlsx')
        sheet = wb.create_sheet('sheet.xlsx', index=0)
        # 添加一行
        row = ['添加一行']
        sheet.append(row)  # 这一步就可以做到将12345插入到一行中。
        wb.save('sheet.xlsx')  # 保存文件，注意以xlsx为文件扩展名
        print('文件保存成功')

    def duquwenjian(self):
        wb = openpyxl.load_workbook('sheet.xlsx')
        sheet = wb['sheet.xlsx']

        for row_idx in range(1, sheet.max_row + 1):
            print(sheet['A{}'.format(row_idx)].value)
            # 只能读取一列


e = ExcelManual()
e.found_file()  # 创建建表格
e.examine()  # 查看表格
e.duquwenjian()

#
#
# 二、选作题
# 1.编写如下程序
# 编写一个工具箱类，需要有工具的名称、功能描述、价格，
# 能够添加工具、删除工具、查看工具，并且能获取工具箱中工具的总数。
