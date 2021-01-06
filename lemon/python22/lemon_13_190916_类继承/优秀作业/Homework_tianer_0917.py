#!/usr/bin/env python3
# -*- coding:utf-8 -*-
# @time     :2019/9/17 20:54
# @author   :天儿
# @file     :Homework_tianer_0917.py
# @software :win10 python3.7.2

# 1、在之前定义的手机类下面定义智能手机类和苹果手机类。
#   智能手机打电话具有录音功能。
#   苹果手机打电话不仅具有录音功能，还有 facetime 功能

# _________________普通手机_________________


class MobilePhone(object):

    def __init__(self, band, model, color):
        """ 初始化手机的品牌、型号、颜色"""
        self.band = band
        self.model = model
        self.color = color

    def call(self, phone_number):
        """父类打电话"""
        return(f"呼叫电话号码{phone_number}")

    def send_message(self,phone_number,content):
        """父类发短信"""
        return(f"给{phone_number}发送短信{content}")

    def listen_music(self, music_name, *args):
        """父类听音乐"""
        return(f"正在循环播放以下歌曲：{music_name}{args}")

# _________________智能手机_________________


class SmartPhone(MobilePhone):

    def call(self, phone_number, if_record):  # 重写手机父类的call,并添加参数if_record
        """子类 打电话-重写，并根据参数判断是否录音"""
        if if_record == "录音":
            return (f"呼叫电话号码{phone_number},并录音")
        else:
            return super().call(phone_number)  # 超继承手机父类

# _________________iphone手机_________________

class IPhone(SmartPhone):

    band = "apple"

    def __init__(self, model, color):
        """重写__init__ 函数"""
        super().__init__(self.band, model, color)   # 超继承


    def call(self, phone_number, if_record, if_facetime):  # 重写# 重写智能手机父类的call,并添加参数if_facetime
        """重写call"""
        if if_record == "录音":
            return (f"呼叫电话号码{phone_number},开启录音功能")

        elif if_facetime == "视频":
            return (f"呼叫电话号码{phone_number},开启视频功能")

        elif if_record == "录音" and if_facetime == "视频":
            return (f"呼叫电话号码{phone_number},开启 录音+视频 功能")

        else:
            return super().call(phone_number,"不录音")   # 超继承智能手机父类的call


smart_phone = SmartPhone('ViVo', 'X20', "红")
print(smart_phone.call(18866666666, '录音'))

iphone7 = IPhone("7plus", "银白")
print(iphone7.call(18800000001, "录音", '视频'))
print(iphone7.call(18800000002, "录音", '不视频'))
print(iphone7.call(18800000003, "不录音", '视频'))
print(iphone7.call(18800000004, '不录音', "不视频"))




# 2、定义一个 ExcelManual 类。具有获取 sheet 表单， 读取单元格 和 修改单元格功能。
from openpyxl import *

# from openpyxl import load_workbook
# # # load_workbook 用来读取excle表里面的数据

# from openpyxl import workbook
# # # workbook 用来新建工作簿以及写入数据到excle中

# filename = r"C:\data\test.xlsx"
# wb = load_workbook(filename)
# wh = wb["Sheet1"]
# ws = wb.active
# wh.cell(3, 4).value = "13123"
# ws['A1'] = '修改的内容'
# wb.save(filename) # 保存内容

# ——————————答题如下————————————

class ExcelManual(object):

    def __init__(self, file_name, sheet_name):
        """初始化需要使用的 excle 文件和需要使用的一个sheet表"""
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_sheet(self):
        """获取表单"""
        wb = load_workbook(self.file_name)  # 打开名为self.file_name的excle文件
        sheet = wb[self.sheet_name]    # 获取名为self.sheet_name的sheet表
        return wb,sheet

    def get_cell(self, row, line):
        cell_value = self.get_sheet()[1].cell(row, line).value  # 找到第row行，第line列的数据，并读取值
        return cell_value

    def modify_cell_test1(self, row, line, new_value):  # 修改方式1，使用坐标
        # wb = self.get_sheet()[0]
        # ws = self.get_sheet()[1]  这样子写是不对的，相当于重复打开了文件，wb和ws打开的不是一个文件

        # wb = load_workbook(self.file_name)
        # ws = self.get_sheet()[1]  这样子写也是不对的，也是相当于重复打开了文件，wb和ws打开的不是一个文件

        wb,ws = self.get_sheet()  # 这样写两者打开的是同一个文件
        ws.cell(row, line).value = new_value  # 修改第row行，第line列的数据 为new_value
        wb.save(self.file_name)  # 修改后关闭
        return new_value

    def modify_cell_test2(self, cell, new_value):   # 修改方式2，使用单元格名称
        wb = self.get_sheet()[0]
        ws = wb.active
        ws[cell] = new_value  # 修改任意单元格，例如"A1" 为cell_new_value
        wb.save(self.file_name)
        return new_value


wb = ExcelManual(r"C:\data\test.xlsx", "Sheet1")   # 实例化一个对象，已知本地该处有已存在的表格和sheet表
print(wb.get_sheet()[1])                           # 打开文件获取到sheet1表         # <Worksheet "Sheet1">
print(wb.get_cell(3, 4))                           # 获取到 第3列，第4行的值        # 34
print(wb.modify_cell_test1(5, 6, '88888'))         # 更改 第5列，第6行 的值为88888  # 88888
print(wb.get_cell(5, 6))                           # 更改后再次查看第5列，第6行 的值     # 88888
print(wb.modify_cell_test2('c1', '你好'))           # 更改C1            # '你好'
print(wb.get_cell(1, 3))                           # 更改后再次查看C1的值 # '你好'





# 二、选作题
# 1.编写如下程序
# 编写一个工具箱类，需要有工具的名称、功能描述、价格，能够添加工具、删除工具、查看工具，并且能获取工具箱中工具的总数。


class ToolCase(object):

    case = []    # 存放工具
    position = 50  # 初始工具箱有50个位置
    number = 0   # 初始有0个工具

    def add(self, name, description, price):
        """放入工具"""
        if self.position > 0:  # 有位置
            self.case.append({"name": name, "description": description, "price": price})
            self.number += 1  # 存放后，数量+1
            print(f"工具箱中放入了{name}工具！")
        else:
            print("工具箱放满了")
        return name

    def delete(self, name):
        """根据工具名字取出某个工具"""
        for i in self.case:
            if i['name'] == name:
                self.case.remove(i)
                self.number -= 1   # 取出后数量-1
                print(f"取出了工具{name}")
                break
        print(f"工具箱中已没有{name}工具")
        return name

    def check_content(self):
        """打印工具箱的剩余全部工具"""
        for i in self.case:
            print(str(i))
        return self.case

    def check_num(self):
        """查看当前有多少件工具"""
        print(f"当前工具箱中一共放了{self.number}个工具")
        return self.number





a = ToolCase()
a.add("锯", "锯木头", "20元")
a.add("锤子", "钉钉子", "5元")
a.add("电钻", "钻墙", "3元")
a.delete('卡尺')
a.delete('锤子')
a.check_content()
a.check_num()
