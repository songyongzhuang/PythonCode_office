# -*- coding: utf-8 -*- 
# @Time     : 2019/8/17 23:49 
# @Author   : wolf_eye 
# @Email    : 15840995236@163.com 
# @File     : homework_0916.py
# @student  : 狼眸

import openpyxl

# Q1:在之前定义的手机类下面定义智能手机类和苹果手机类。智能手机打电话具有录音功能。
# 苹果手机打电话不仅具有录音功能，还有 facetime 功能


class MobilePhone(object):
    def __init__(self, system_type, logo, size, version):
        self.sys_type = system_type     # 手机系统类型
        self.logo = logo                # 手机品牌
        self.size = size                # 手机型号
        self.version = version          # 手机版本
        print(f'手机系统类型{self.sys_type}, 手机品牌{self.logo}, 手机型号{self.size}, 机版本{self.version}')

    def call_number(self, name, phone_number):
        print(f'呼叫{name},电话号码{phone_number}')

    def send_message(self, content):
        return f'发送内容：{content}'

    def listen_music(self, music_name):
        return f'正在播放歌曲：{music_name}'


class IntelligentPhone(MobilePhone):
    """智能手机类：继承手机基类"""
    is_intelligent_phone = '我是智能手机'

    # 超继承
    # def call(self, name, phone_number):
    #     super().call_number(name, phone_number)

    def call_number(self, name, phone_number):
        """重写父类方法，打电话方法"""
        print(f'{self.is_intelligent_phone},我正在呼叫{name},电话号码{phone_number}')

    def sound_recording(self):
        """子类定义新方法，录音方法"""
        print(f'{self.is_intelligent_phone},我有录音功能')


class IPhone(IntelligentPhone):
    """苹果手机类，继承智能手机类"""
    logo = 'iphone'

    def __init__(self, system_type, size, version):
        super().__init__(system_type, self.logo, size, version)
        print(f'手机系统类型{self.sys_type}, 手机品牌{self.logo}, 手机型号{self.size}, 机版本{self.version}')

    def facetime(self):
        """定义视频通话方法"""
        print(f'我是{self.logo}手机，我有视频通过功能')


# Q2：定义一个 ExcelManual 类。具有获取 sheet 表单， 读取单元格 和 修改单元格功能
class ExcelManual(object):
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name


    def read_form(self):
        """读取表单数据"""
        excel = openpyxl.load_workbook(self.file_path)
        sheet = excel[self.sheet_name]
        read_data = []
        # 按行读取，每行内容存在列表中，再将行列表存到表单列表中
        for i in range(1, sheet.max_row + 1):
            read_line = []
            for j in range(1, sheet.max_column + 1):
                read_line.append(sheet.cell(i, j).value)
            read_data.append(read_line)
        return read_data

    def edit_form(self, row, column, update_content):
        """修改表单方法"""
        excel = openpyxl.load_workbook(self.file_path)
        sheet = excel[self.sheet_name]
        sheet.cell(row, column).value = update_content
        excel.save(self.file_path)
        return sheet.cell(row, column).value




# Q3：编写一个工具箱类，需要有工具的名称、功能描述、价格，能够添加工具、删除工具、查看工具，并且能获取工具箱中工具的总数。
# 方法一:工具箱使用列表存储
class ToolBox:

    def __init__(self, tool_list):
        """初始化实例对象列表，本程序采用列表存储工具箱工具"""
        self.tool_list = tool_list

#    def addtool(self, name, descripe, price):
    def addtool(self, *args):
        """添加工具"""
        toolbox_dict = {}
        toolbox_dict['name'] = args[0]
        toolbox_dict['descripe'] = args[1]
        toolbox_dict['price'] = args[2]
        self.tool_list.append(toolbox_dict)
        return self.tool_list

    def deltool(self, del_index):
        """删除工具"""
        return self.tool_list.pop(del_index)

    def looktool(self):
        """查看工具"""
        for element in self.tool_list:
            print(f'{self.tool_list.index(element)}:{element}')

# 方法二:工具使用文件存储
class ToolBoxTwo():
    def __init__(self, file):
        self.file = file

    def addtool(self, write_data):
        """增加工具，写入文件"""
        with open(self.file, 'a', encoding='UTF-8') as f:
            f.writelines(write_data)


    def looktool(self):
        with open(self.file, 'r+', encoding='UTF-8') as f:
            for file_data in f.readlines():
                print(file_data, end='')
            f.seek(0)
            return f.readlines()

    def deltool(self, line):
        """删除工具"""
        del_data = self.looktool()
        del_data.pop(line)
        with open(self.file, 'w+', encoding='UTF-8') as f:
            self.addtool(del_data)
            return self.file



if __name__ == '__main__':
# Q1: Test program
    # 实例化手机类对象MobilePhone
    mobile_phone = MobilePhone('IOS', 'Apple', '6.0', '10.0')
    mobile_phone.call_number('wolf', '15800000000')
    message = mobile_phone.send_message('Python is my life')
    print(message)
    print(mobile_phone.listen_music('芒种'))

    # 实例化手机类 子类智能手机类对象IntelligentPhone
    intelligent_phone = IntelligentPhone('IOS', 'Apple', '6.1', '11.0')
    # 调用子类重写方法
    intelligent_phone.call_number('wolf', '15811111111')
    # 调用父类听音乐方法
    print(intelligent_phone.listen_music('攀登'))
    # 调用子类新定义录音方法
    intelligent_phone.sound_recording()

    # 实例化 智能手机类子类 苹果手机类 对象
    iphone = IPhone('IOS', '6.2', '12.0')
    # 调用子类新定义视频通话方法
    iphone.facetime()
    # 调用父类打电话方法
    iphone.call_number('wolf', '15822222222')
    # 调用父类的父类 听音乐 方法
    print(iphone.listen_music('不服'))


# Q2: Test program
    filepath = 'D:\\test.xlsx'
    # 创建excel
    ex = openpyxl.Workbook(filepath)
    # 创建sheet
    sh_new = ex.create_sheet('Sheet1')
    # 保存创建的excel、sheet
    ex.save(filepath)
    # 再刚创建的sheet中更新测试数据
    excel = ExcelManual(filepath, 'Sheet1')
    test_list = [[1, 2, 3], ['wolf', 'wenwen', 'chaoge']]
    for row in range(1, 3):
        for column in range(1, 4):
            data = excel.edit_form(row, column, test_list[row - 1][column - 1])
            print(f'表单行{row}，表单列{column}:数据{data}更新成功')
    # 读取excel表单数据
    read_data = excel.read_form()
    print('表单数据为', read_data)


# Q3: Test program
    # 实例化对象
    toolbox = ToolBox([])
    # 添加工具，测试数据
    tool = [['锤子', '砸东西', '100'], ['剪刀', '剪东西', '20']]
    for i in tool:
        toolbox.addtool(*i)
    # 查看工具
    toolbox.looktool()
    # 删除工具
    toolbox.deltool(0)
    toolbox.looktool()

# Q3: 方法二：Test program
file_txt_path = 'D:\\test.txt'
tool = ["'锤子', '砸东西', '100'\n", "'剪刀', '剪东西', '20'\n"]
# 实例化对象
toolboxtwo = ToolBoxTwo(file_txt_path)
# 写入tool中测试数据
toolboxtwo.addtool(tool)
# 查看工具
print(toolboxtwo.looktool())
# 删除工具
toolboxtwo.deltool(1)
print(toolboxtwo.looktool())