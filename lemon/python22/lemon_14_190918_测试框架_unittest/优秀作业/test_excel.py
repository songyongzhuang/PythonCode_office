from openpyxl import load_workbook


class DoExcel:

    def do_excel(self, case_num, title_num, excel_name, sheet_name):
        wb = load_workbook(excel_name)
        sheet = wb[sheet_name]
        test_data = []
        for i in range(2, case_num+2):
            Sub_dict = {}
            for j in range(1, title_num+1):
                Sub_dict[sheet.cell(1, j).value] = sheet.cell(i, j).value
            test_data.append(Sub_dict)
        return test_data

# a = DoExcel()
# b = a.do_excel(8,5,'testdata.xlsx','name')
# print(b)
    def write_back(self, excel_name, sheet_name, row, testresult, ifpass):
        wb = load_workbook(excel_name)
        sheet = wb[sheet_name]
        sheet.cell(row, 6).value = testresult
        sheet.cell(row, 7).value = ifpass
        wb.save(excel_name)
