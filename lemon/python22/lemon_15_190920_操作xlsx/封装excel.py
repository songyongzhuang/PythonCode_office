# --*-- coding : utf-8 --*--
# Project      : python22
# Current file : 封装excel.py
# Author       : 大壮
# Create time  : 2020-01-11 17:54
# IDE          : PyCharm
# TODO 成长很苦，进步很甜，加油！
import openpyxl


class ExcelHandle(object):
    """ excel封装 """

    def __init__(self):
        """ 也可以把文件名和表名写到init里面 """
        pass

    def select_form(self, file_name, sheet_name):
        """
        选择表单，判断传入的 sheet_name 是整数根据索引，不是整数根据表单名处理
        :param file_name: excel文件路径
        :param sheet_name: 表单的名字或者索引
        """
        wb = openpyxl.load_workbook(file_name)
        if isinstance(sheet_name, int):
            try:
                return wb.worksheets[sheet_name]
            except IndexError:
                print('输入的表单索引有误')
        elif isinstance(sheet_name, str):
            try:
                return wb[sheet_name]
            except KeyError:
                print('输入的表单名称有误')
        else:
            print('输入有误，请输入表单索引或表单名称')

    def read_one_data(self, file_name, sheet_name, row, column):
        """
        读取 一个单元格的数据 , 根据第几行第几个获取数据
        :param file_name: excel文件路径
        :param sheet_name: 表单的名字或者索引
        :param row: 行
        :param column: 列
        :return:
        """
        sheet = self.select_form(file_name, sheet_name)
        return sheet.cell(row, column).value

    def read_line(self, file_name, sheet_name, line):
        """
        读取 一列单元格的数据
        :param file_name: excel文件路径
        :param sheet_name: 表单的名字或者索引
        :param line:
        """
        # 调用上面选择表单
        sheet = self.select_form(file_name, sheet_name)
        # 获取单元格
        sheet_data = sheet[line]
        # 元组 （Cell(1,1), Cell(1,2）
        data = []
        for c in sheet_data:
            data.append(c.value)
        return data

    def read_all(self, file_name, sheet_name, start_row=2, start_column=1):
        """
        读取所有的数据,
        :param file_name: excel文件路径
        :param sheet_name: 表单的名字或者索引
        :param start_row: 从第几'行'开始
        :param start_column: 从第几'列'开始
        :return:
        """
        sheet = self.select_form(file_name, sheet_name)
        # 使用 列表推导式  列表中的索引开始是从 1 开始, 而不是 0
        header = [colmn.value for colmn in sheet[1]]
        data = []
        # sheet.max_row 获取最大行数
        for row in range(start_row, sheet.max_row + 1):
            row_data = []
            # sheet.max_column 获取最大列数
            for column in range(start_column, sheet.max_column + 1):
                row_data.append(sheet.cell(row, column).value)
            # 把标题头和输入组合成字典 把两个列表组合成字典
            row_data = dict(zip(header, row_data))
            data.append(row_data)
        return data

    @staticmethod
    def write(file_name, sheet_name, row, column, data):
        """
        保存数据，使用 静态方法 一次修改一个
        :param file_name: excel文件路径
        :param sheet_name: 表单的名字
        :param row: 修改数据的'行'
        :param column: 修改数据的'列'
        :param data: 修改的内容
        :return:
        """
        wb = openpyxl.load_workbook(file_name)
        # todo 使用 get_sheet_by_name 会告诉你这个函数被弃用了
        #  sheet = wb.get_sheet_by_name(sheet_name)
        # 建议使用
        sheet = wb[sheet_name]
        sheet.cell(row, column).value = data
        # 保存关闭
        wb.save(file_name)
        wb.close()


if __name__ == '__main__':
    eh = ExcelHandle()
    # 获取表单
    print(eh.select_form(r'D:\data.xlsx', 0))
    # 一个数据
    print(eh.read_one_data(r'D:\data.xlsx', 0, 1, 1))
    # 获取所有数据
    print(eh.read_all(r'D:\data.xlsx', 0))
    # 保存数据
    eh.write(r'D:\data.xlsx', 'Sheet1', 1, 1, 'URL')
