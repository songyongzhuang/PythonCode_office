#!/usr/bin/env python3
# -*- coding:utf-8 -*-
# datetime:2019/9/20 20:45
# email: wagyu2016@163.com
# author: 雨泽
# copyright: 湖南省零檬信息技术有限公司

# 操作 Excel ，使用第三方库，别人写好大的代码？
# 别人有现成的库，尽量不要自己写，重复造轮子。
# python: 轮子非常，随你使用。
# 拼自己的智慧整合。

# openpyxl 操作 2010 版以后的 xlsx,
# 2003 经典， xlrd, xlwt
# tablib, 都支持。

# pip install openpyxl

import openpyxl
# 打开excel 文件，获取工作簿 workbook
# 2，打开表单
# 3，获取单元格信息

# 1，获取一个单元格， 字符串
# 2，读取一列/一行， 列表，元组，字典
# 3， 获取多行/多列。   嵌套的列表，

# 获取工作簿
wb = openpyxl.load_workbook(r'D:\demo.xlsx')
print(wb)

# 获取所有表单
sheet = wb.worksheets[2]
# print(sheet)

# 获取表单
# sheet = wb.get_sheet_by_name('Sheet3')
# 类似于字典的操作
# sheet = wb['Sheet3']

# 三个方法, 激活的
# sheet = wb.active


# 获取单元格的值, 以 1 开头，和python 索引不同
cell = sheet.cell(1,1).value
print(cell)

# 获取一行的数据Q
# a = sheet[1]
# print(a)
#
# # 获取多行的数据, 和列表切面的区别
# b = sheet[1:3]
# print(b)

# 获取一列的数据
# a = sheet["B"]
# print(a)

# 测试过程
# 获取所有的行 ，最简单的方式
# 生成器，

# data = sheet.rows
# data_list = list(data)[1:]
# # [('url','data', 'case_id'),('baidu'...),(...)]
# print(data_list)
#
# # 嵌套列表的作业重新做一遍
# new_data = []
# for row in data_list:
#     # (<Cell 'Sheet3'.A2>, <Cell 'Sheet3'.B2>, <Cell 'Sheet3'.C2>)
#     row_data = []
#     for cell in row:
#         # <Cell 'Sheet3'.A2>
#         row_data.append(cell.value)
#     new_data.append(row_data)
# print(new_data)


# 第二种方式操作
# 获取最大行数
# datas = []
# max_row = sheet.max_row
# max_column = sheet.max_column
#
# for row in range(2, max_row+1):
#     row_data = []
#     for column in range(1, max_column+1):
#         row_data.append(sheet.cell(row, column).value)
#     datas.append(row_data)
# print(datas)

# 读取所有的数据 cell
# [()], [{}]

# 写
# 修改
sheet.cell(1, 1).value = 'new_url'
sheet.cell(1, 2).value = 'new_data'

# 保存
wb.save(r'D:\data.xlsx')
# 关闭
wb.close()

# 数据复用
# 配置文件
