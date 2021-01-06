# --*-- coding : utf-8 --*--
# Project      : python_lemon_作业
# Current file : lemon_自个练习.py
# Author       : 大壮
# Create time  : 2019-09-21 11:06
# IDE          : PyCharm
# TODO 成长很苦，进步很甜，加油！
import openpyxl

# 1.打开excel文件
# 2.获取工作簿 workbook
# 3.获取单元格信息
#   3.1.获取一个单元格，
#   3.2.获取一列/一行单元格，用容器装载
#   3.3.获取多列/多行单元格，是嵌套的容器装载

# ------------------------------- 打开excel文件 -----------------------------------

# 打开excel文件
wb = openpyxl.load_workbook(r'D:\data.xlsx')  # 需要传递文件路径或者文件对象
print(wb)

# ---------------------------------- 表单 --------------------------------

# 获取所有表单
# sheet = wb.worksheets
# print(sheet)

# 根据索引获取表单
sheet = wb.worksheets[1]
print(sheet)


# 根据名字获取表单
sheet1 = wb.get_sheet_by_name('Sheet3')
sheet = wb['Sheet3']
print(sheet1)
print(sheet)

# # active：当前的、默认的、被激活的
sheet = wb.active
print(sheet)

# ----------------------------------- 单元格 -------------------------------

# 获取单元格(cell)的值, 文件的单元格是以’1‘ 开头，不是零
a = eval('1')
print(a)
cell = sheet.cell(1, 1).value
print(cell)

# 获取多行数据
a = sheet[1:2]
print(a)

# 获取一列数据
a = sheet['A']
print(a)


# 第一种方式操作
# 获取所有行，最简单方式 rows
# generator 生成器，可以一行一行的提取数据 rows
# 生成器可以转换为列表
data = sheet.rows
data_list = list(data)[1:]  # [1:] 是为了去掉文件的标题, 只取数据
# [(第一行数据), (第二行数据), ...]
print(data_list)

# 循环遍历每个单元格，value返回的所有值
new_data = []
for row in data_list:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    new_data.append(row_data)
print(new_data)


# 第二种方式操作
# 获取最大行数
datas = []  # 创建空列表
max_row = sheet.max_row  # 获取最大行数
max_column = sheet.max_column  # 获取最大列数
# range 自动生成类似于列表的
for row in range(2, max_row+1):  # 2 为了去掉文件的标题, 只取数据
    row_data = []
    for column in range(1, max_column+1):  # 表格是从1开始的，要和索引区分
        row_data.append(sheet.cell(row, column).value)
        # 获取单元格(cell)--row 单元格的行--column 单元格的列--value 把单元格的值提取出来
    datas.append(row_data)
print(datas)


# 写操作(修改)
sheet.cell(1, 1).value = '888888'
# 保存
wb.save(r'D:\data.xlsx')
# 关闭
wb.close()
