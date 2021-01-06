# --*-- coding : utf-8 --*--
# Project      : python_lemon_作业
# Current file : test_测试相加.py
# Author       : 大壮
# Create time  : 2019-09-22 17:26
# IDE          : PyCharm
# TODO 成长很苦，进步很甜，加油！


import unittest
import openpyxl

# 打开excel文件
wb = openpyxl.load_workbook(r'D:\data.xlsx')  # 需要传递文件路径或者文件对象
sheet = wb.worksheets[0]  # 打开表单

datas = []  # 创建空列表
max_row = sheet.max_row  # 获取最大行数
max_column = sheet.max_column  # 获取最大列数
# range 自动生成类似于列表的
for row in range(2, max_row+1):  # 2 为了去掉文件的标题, 只取数据
    row_data = []
    for column in range(3, max_column+1):  # 表格是从1开始的，要和索引区分
        row_data.append(sheet.cell(row, column).value)
        # 获取单元格(cell)--row 单元格的行--column 单元格的列--value 把单元格的值提取出来
    datas.append(row_data)


def add(a, b):  # add 加起来
    """ 相加 """''
    return a + b


class TestExcel(unittest.TestCase):

    def test_add_success(self):
        """ 判断表达式是否为真 """''
        self.assertTrue(datas[0][2] == add(datas[0][0], datas[0][1]))

    def test_add_error(self):
        """如果确定两个对象不相等，则失败。"""''
        self.assertEqual(datas[1][2], add(datas[1][0], datas[1][1]))


if __name__ == '__main__':

    unittest.main()
