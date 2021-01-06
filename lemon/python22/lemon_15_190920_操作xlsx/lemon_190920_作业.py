# --*-- coding : utf-8 --*--
# Project      : python_lemon_作业
# Current file : lemon_190920_作业.py
# Author       : 大壮
# Create time  : 2019-09-20 22:14
# IDE          : PyCharm
# TODO 成长很苦，进步很甜，加油！
import openpyxl
# 第一：excel类封装需要提供以下功能：
# 1、选择表单功能
# 2、读取一个单元格的数据功能
# 3、读取一行数据  功能
# 4、读取表单中所有数据功能
# 5、往单元格中写入数据功能
# 6、保存数据功能


class ExcelManual(object):

    def __init__(self, file, ):
        self.file = file

    def open_file(self):
        """ 打开文件 """''
        wb = openpyxl.load_workbook(self.file)
        return wb

    def get_sheet(self, indexes):
        """ 选择表单功能 根据索引 """''
        sheet = self.open_file().worksheets[indexes]
        return sheet

    def get_sheet_two(self, indexes):
        """ 选择表单功能 根据名字获取表单，使用这个没有提示... """
        sheet = self.open_file()[indexes]
        return sheet

    def one_cell(self, indexes, coordinate):
        """ 读取一个单元格的数据功能
            indexes 传表单索引
            coordinate 坐标轴
        """''
        b = []
        for i in coordinate:
            b.append(i)
        cell = self.get_sheet_two(indexes).cell(eval(b[0]), eval(b[-1])).value
        return cell

    def row_cell(self, indexes, coordinate):
        """ 读取一行数据 功能
        """''
        cell = self.get_sheet(indexes)[coordinate]
        b = []
        for i in cell:
            b.append(i.value)
        return b

    def all_data(self, indexes):
        """ 读取表单中所有数据功能 
            indexes 表单索引
        """''
        data = self.get_sheet(indexes).rows
        data_list = list(data)  #
        # 循环遍历数据
        a = []
        for i in data_list:
            b = []
            for j in i:
                b.append(j.value)
            a.append(b)
        return a

    def all_data_two(self, indexes, row, column):
        """ 读取表单中所有数据功能
            indexes 表单索引
            hang 行
            lie 列
        """''
        # 获取最大行数
        max_row = self.get_sheet(indexes).max_row
        # 获取最大列数
        max_column = self.get_sheet(indexes).max_column
        a = []
        for i in range(row, max_row+1):  # 行
            b = []
            for j in range(column, max_column+1):  # 列
                b.append(self.get_sheet(indexes).cell(i, j).value)
            a.append(b)
        return a

    def write_file(self, indexes, coordinate, revised):
        """ 往单元格中写入数据功能
            indexes 表单索引
            coordinate 修改数据的坐标
            revised  修改的数据
        """
        b = []
        for i in coordinate:
            b.append(i)
        wb = openpyxl.load_workbook(self.file)  # 文件路径和名称
        sheet = wb.worksheets[indexes]  # 表单
        sheet.cell(eval(b[0]), eval(b[-1])).value = revised  # 坐标和数据
        wb.save(self.file)  # 保存
        return revised

    def save_file(self):
        """ 保存文件, 关闭文件 """''
        self.open_file().save(self.file)
        self.open_file().close()


xlsx = ExcelManual(r'D:\data.xlsx')
print('根据索引选择表单', xlsx.get_sheet(2))  # 根据索引选择表单
# print(xlsx.get_sheet_two('Sheet3'))  # 根据名字获取表单

# 读取一个单元格的数据功能 第一个参数：表单名字 第二个参数：单个单元格的坐标
print('读取一个单元格的数据功能', xlsx.one_cell('Sheet3', '1,2'))

# 读取一行单元格的数据功能 第一个参数：表单索引 第二个参数：第几行
print('读取一行数据 功能', xlsx.row_cell(2, 1))

# 读取表单中所有数据功能 参数：表单索引
print('表单中所有数据:方法一', xlsx.all_data(2))

# 读取表单中所有数据功能 参数：表单索引
# 第一个参数表单的索引, 第二个：行, 第三个：列
print('表单中所有数据456:方法二', xlsx.all_data_two(indexes=0, row=2, column=2))

# 往单元格中写入数据功能 表单索引 修改数据的坐标 修改的数据
# print(f'保存数据为：', xlsx.write_file(1, '1,1', '修改的数据'))

# 保存文件, 关闭文件  一定确保文件没有在其他地方打开，否则报错
# xlsx.save_file()

'''
# 第二：请设计测试数据，对封装的excel类功能进行测试。
# excel中的具体内容，由各位同学自由发挥
# 导入HTMLTestRunnerNew
from HTMLTestRunnerNew import HTMLTestRunner  # 生成HTML文件调用模块
import os  # 使用os模块获取路径
import unittest  # python自带单元测试模块

# 初始化 loader  测试加载器（TestLoader）
loader = unittest.TestLoader()
start_dir = os.path.dirname(os.path.abspath(__file__))
suite = loader.discover(start_dir)  # 找到文件路径运行

with open('测试报告_001.html', 'wb') as f:
    """ stream 文件流，可以传文件    verbosity 详细程度
    title=文件标题  description=文件的注释   tester=测试人员 """
    runner = HTMLTestRunner(f,
                            verbosity=2,
                            title='我是文件标题',
                            tester='我是测试人员',
                            description='我是文件注释')
    # 运行
    runner.run(suite)
'''
