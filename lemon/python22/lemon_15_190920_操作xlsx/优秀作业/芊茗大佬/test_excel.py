'''第一：excel类封装需要提供以下功能：

1、选择表单功能

2、读取一个单元格的数据功能

3、读取一行数据  功能

4、读取表单中所有数据功能

5、往单元格中写入数据功能

6、保存数据功能'''
import openpyxl
import unittest


def add(a,b):
    return a + b


class Excel:

    def __init__(self,file_path):
        self.file_path = file_path


    def choose_sheet(self,sheetname):
        '''选择表单功能'''
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb[sheetname]
        return sheet

    def get_cell(self,sheetname,row,column):
        '''读取一个单元格的数据功能'''
        cell = self.choose_sheet(sheetname).cell(row,column).value
        return cell

    def get_row_cell(self,sheetname,row):
        '''读取一行数据功能'''
        max_column = self.choose_sheet(sheetname).max_column
        list_1 = []
        for column in range(1,max_column+1):
            cell = self.choose_sheet(sheetname).cell(row,column).value
            list_1.append(cell)
        return list_1


    def get_all_cell(self,sheetname):
        '''读取表单中所有数据功能'''
        data =  self.choose_sheet(sheetname).rows
        data_list = list(data)[1:]
        new_data = []
        for row in data_list:
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            new_data.append(row_data)
        return new_data


    def write_cell(self,sheetname,row,column,value):
        '''往单元格中写入数据功能'''
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb[sheetname]
        sheet.cell(row,column).value = value
        wb.save(self.file_path)
        wb.close()
        return value

    def excel_save(self):
        '''保存数据功能'''
        wb = openpyxl.load_workbook(self.file_path)
        wb.save(self.file_path)
        wb.close()


excel = Excel(r"D:\test.xlsx")

'''自己调试部分'''
# print(excel.choose_sheet("Sheet1"))
# print(excel.get_cell("Sheet1",1,1))
# print(excel.get_row_cell("Sheet1",3))
# print(excel.get_all_cell("Sheet1"))
# print(excel.write_cell("Sheet1",3,1,"随便写"))

class TestAdd(unittest.TestCase):


    def setUp(self):
        '''前置条件。
        测试用例方法执行之前自动运行setUp里面的程序
        '''
        print("用例执行前置条件")

    def tearDown(self):
        '''后置条件。
        测试用例执行方法之后自动运行tearDown里面的程序
        '''
        print("用例执行后置条件")

    def test_1_success(self):     # 按照定义的加法方法获取第2行前两个单元格的内容，与该行第三个单元格内容比较
        self.assertTrue(add(excel.get_cell("Sheet1",2,1),excel.get_cell("Sheet1",2,2)) == excel.get_cell("Sheet1",2,3))

    def test_1_error(self):        # 按照定义的加法方法获取第2行前两个单元格的内容，与自定义的一个错误的数字进行比较
        self.assertTrue(add(excel.get_cell("Sheet1",2,1),excel.get_cell("Sheet1",2,2)) == 4)

    def test_2_success(self):      # 获取正确的Sheet1表单第三行的内容
        self.assertEqual(excel.get_row_cell("Sheet1",3),[5,6,10])

    def test_2_error(self):       # 获取错误的Sheet1表单第三行的内容
        self.assertEqual(excel.get_row_cell("Sheet1",3),[12,13,25])

    def test_3_success(self):     # 单元格修改数据后再次获取该单元格的值
        self.assertTrue(excel.write_cell("Sheet1",4,1,"修改1") == excel.get_cell("Sheet1",4,1))

    def test_3_error(self):        # 单元格修改数据后，获取对应错误单元格的值
        self.assertTrue(excel.write_cell("Sheet1",4,2,"修改2") == excel.get_cell("Sheet1",4,1))