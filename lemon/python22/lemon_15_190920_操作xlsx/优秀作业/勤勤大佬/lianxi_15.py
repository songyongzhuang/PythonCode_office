# @Time : 2019/9/21 20:42 
# @Author : ZhangHaiQin
# @File : lianxi_15.py 
# @Software: PyCharm
import openpyxl
from openpyxl import load_workbook

"""
第一：excel类封装需要提供以下功能：
1、选择表单功能
2、读取一个单元格的数据功能
3、读取一行数据  功能
4、读取表单中所有数据功能
5、往单元格中写入数据功能
6、保存数据功能
"""


class ExcelFun:
    def __init__(self, path):
        # 文件名与选择的表单
        self.path = path

    # 选择工作簿
    def choice_wb(self):
        wb = openpyxl.load_workbook(self.path)
        return wb

    # 选择表单功能
    def choice_sheet(self, sheet):
        wb = self.choice_wb()
        sheet1 = wb[sheet]
        print(sheet1)
        return sheet1

    # 读取一个单元格的数据功能
    def data_cell(self, sheet, row, column):
        a = self.choice_sheet(sheet).cell(row, column)
        b = a.value
        print(b)
        return b

    # 读取一行数据  功能
    def data_row(self, sheet, row):
        sheet3 = self.choice_sheet(sheet)
        a = sheet3[row]
        b = []
        for cell in a:
            b.append(cell.value)
        print(b)
        return b

    # 读取表单中所有数据功能
    def data_all(self, sheet):
        sheet4 = self.choice_sheet(sheet)
        data = sheet4.rows
        list_data = list(data)[1:]
        datas = []
        for row in list_data:
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            datas.append(row_data)
        print(datas)

    # 往单元格中写入数据功能
    def data_write(self, sheet, row, column, data):
        wb = self.choice_wb()
        sheet5 = wb[sheet]
        sheet5.cell(row, column).value = data
        print(data)
        self.data_save(wb)

    # 保存数据功能
    def data_save(self, wb):
        wb.save(self.path)
        wb.close()


# excel_fun = ExcelFun("C:\柠檬班/demo.xlsx")
# excel_fun.choice_sheet("Sheet3")
# excel_fun.data_cell("Sheet3", 1, 1)
# excel_fun.data_row("Sheet3", 1)
# excel_fun.data_all("Sheet3")
# excel_fun.data_write("Sheet3", 5, 5, "你好")

"""
第二：请设计测试数据，对封装的excel类功能进行测试。
excel中的具体内容，由各位同学自由发挥
"""
