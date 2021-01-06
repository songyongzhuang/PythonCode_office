# --*-- coding : utf-8 --*--
# Project      : python_lemon_作业
# Current file : 练习_two.py
# Author       : 大壮
# Create time  : 2019-09-22 14:22
# IDE          : PyCharm
# TODO 成长很苦，进步很甜，加油！
import openpyxl
from lemon_15_190920_操作xlsx.封装excel import ExcelHandle
# 打开excel文件
"""
wb = openpyxl.load_workbook(r'D:\data.xlsx')  # 需要传递文件路径或者文件对象
print(wb)  # 打开表单

# # 获取所有表单
# sheet = wb.worksheets
# print(sheet)

# 根据索引获取所有表单
sheet = wb.worksheets[2]
print(sheet)


# 获取单元格
# 获取单元格(cell)的值, 文件的单元格是以’1‘ 开头，不是零
a = sheet.cell(1, 1)
print(a)
cell = sheet.cell(1, 1).value
print(cell)

# # 获取一行的数据
# a = sheet[1]
# print(a)

# 获取多行的数据
b = sheet[1:3]
print(b)

# 获取一列的数据
c = sheet['A']
print(c)


# 获取所有的行，
# generator 生成器, 会一行一行的提取数据
data = sheet.rows
data_list = list(data)[1:]  # [1:]剔除标题不是测试数据
print(data_list)
new_data = []
for row in data_list:
    row_data = []  # 储存每一行的数据
    for cell in row:
        row_data.append(cell.value)
    new_data.append(row_data)
print(new_data)
"""
"""
# # 第二种方法
# # 获取最大行数
# max_rom = sheet.max_row  # 行;横着的
# max_column = sheet.max_column  # 列;竖着的
# print('行', max_rom)
# print('列', max_column)
# url = []
# for row in range(2,max_rom+1):
#     url1 = []
#     for column in range(1, max_column+1):
#         url1.append(sheet.cell(row, column).value)  # 获取单元格(cell)
#     url.append(url1)
# print(url)


# 写操作
sheet.cell(4, 1).value = '666'

# 保存的时候需要添加路径，路径一致就是保存，不一致就是另存为
wb.save(r'D:\data.xlsx')
# 关闭
wb.close()



datas = []  # 创建空列表
max_row = sheet.max_row  # 获取最大行数
max_column = sheet.max_column  # 获取最大列数
# range 自动生成类似于列表的
for row in range(2, max_row+1):  # 2 为了去掉文件的标题, 只取数据
    row_data = []
    for column in range(3, max_column+1):  # 表格是从1开始的，要和索引区分
        row_data.append(sheet.cell(row, column).value)
        # 获取单元格(cell)--row 单元格的行--column 单元格的列--value 把单元格的值提取出来
    datas.append(row_data)
print(datas)
"""

print(ExcelHandle(r'D:\data.xlsx').read_all(0))
