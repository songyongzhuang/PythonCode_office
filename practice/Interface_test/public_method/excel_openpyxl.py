# --*-- coding : utf-8 --*--
# Project      : Interface_test
# Current file : excel_openpyxl.py
# Author       : 大壮
# Create time  : 2020-07-12 11:15
# IDE          : PyCharm
# TODO 成长很苦，进步很甜，加油！
import openpyxl
# 日志
import logging
# 定义的日志  相当重要的那，要不不记录日志
from practice.Interface_test.public_method import logger
# 项目路径
from practice.Interface_test.public_method.dir_path import DirPath


class ExcelHandler:
    """excel 封装"""

    def choose_sheet(self, file_name, sheet_name):
        """选择表单.
        sheet_name 是整数，根据索引获取。
        如果是字符串，根据名字获取 '20190920'
        """
        logging.info(f'文件路径{file_name}，表单：{sheet_name}')
        wb = openpyxl.load_workbook(file_name)
        try:
            if isinstance(sheet_name, int):
                return wb.worksheets[sheet_name]
            elif isinstance(sheet_name, str):
                return wb[sheet_name]
        except:
            return '工作表不存在'

    def choose_grid(self, file_name, sheet_name, row, column):
        """ 获取单个单元格 """
        sheet = self.choose_sheet(file_name, sheet_name)
        logging.info(f'获取单个单元格，文件路径{file_name}，表单：{sheet_name}，坐标为:{row, column}')
        return sheet.cell(row, column).value

    def choose_title(self, file_name):
        """表单名"""
        wb = openpyxl.load_workbook(file_name)
        return wb.worksheets

    def read(self, file_name, sheet_name, start_row=2, start_column=1):
        """获取所有的数据，字典格式"""
        logging.info(f'获取所有数据，文件路径{file_name}，表单：{sheet_name}')
        sheet = self.choose_sheet(file_name, sheet_name)
        max_rom = sheet.max_row  # 行
        max_column = sheet.max_column  # 列
        header = [colmn.value for colmn in sheet[1]]
        data = []
        for row in range(start_row, max_rom + 1):
            row_data = []
            for column in range(start_column, max_column + 1):
                row_data.append(sheet.cell(row, column).value)
            row_data = dict(zip(header, row_data))
            data.append(row_data)
        return data

    @staticmethod  # 写入表格 静态方法
    def write(file_name, sheet_name, row, column, data):
        logging.info(f'写入文件，文件路径{file_name}，表单：{sheet_name}，坐标{row, column}写入数据{data}')
        global sheet  # 函数中调用全局变量的值
        wb = openpyxl.load_workbook(file_name)  # 文件路径
        try:
            if isinstance(sheet_name, int):
                sheet = wb.worksheets[sheet_name]
            elif isinstance(sheet_name, str):
                sheet = wb[sheet_name]
        except:
            logging.info(f'写入，文件路径{file_name}，表单：{sheet_name}，工作表不存在')
            return '工作表不存在'
        # sheet = wb.worksheets[sheet_name]  # 表单
        sheet.cell(row, column).value = data  # 替换或跟新，坐标和数据
        wb.save(file_name)  # 保存
        wb.close()  # 关闭


# 文件路径
xlsx = f'{DirPath.datas_dir}\cases_excel.xlsx'
if __name__ == '__main__':
    eh = ExcelHandler()
    # 选择表单
    print(eh.choose_sheet(xlsx, 'Sheet1'))
    # 获取单个单元格
    print(eh.choose_grid(xlsx, 0, 1, 1))
    # 表单名
    print(eh.choose_title(xlsx))
    # 获取所有的数据，字典格式
    print(eh.read(xlsx, 0))
    # 写入表格
    eh.write(xlsx, 'Sheet1', 4, 2, '替换数据123')
